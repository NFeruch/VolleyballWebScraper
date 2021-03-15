# -*- coding: UTF-8 -*-
# https://www.upwork.com/jobs/~015ddd538746ecf0d8/
# http://www.usavolleyballclubs.com/volleyballclubdirectory.asp
#
# Find:
# 1. Club/Team name yes
# 2. Sport
# 3. Type
# 4. Contact name = Contact
# 5. State = State
# 6. City or County = City
# 7. Email address = More Info
# 8. Related website

import requests, html, pandas as pd
from random import randint
from time import sleep
from bs4 import BeautifulSoup
from openpyxl import load_workbook, Workbook

# wb = Workbook()
# wb.save('test.xlsx')
wb = load_workbook("test.xlsx")
main_worksheet = wb.worksheets[0]
main_worksheet.append(['Club Name', 'Gender', 'Age Group', 'Contact Name', 'City', 'State', 'Email', 'Website', 'Region'])

for page in range(73, 82):
    index = 0
    print('Page: ' + str(page))
    url = r'http://www.usavolleyballclubs.com/volleyballclubdirectory.asp?' + 'do_search=1&keywords=&page_no={}'.format(
        page)
    current_page = requests.get(url)
    current_page_source = current_page.content
    web_soup = BeautifulSoup(current_page_source, 'lxml')

    table = web_soup.find_all('table', {'class': 'table table-striped'})
    trs = str(table[0]).split('</tr>')
    trs.pop()
    for tr in trs:
        tr = html.unescape(tr)
        try:
            name = tr[tr.index('<h3>') + 4:tr.index('</h3>')]
            if '<a' in name:
                name = name[name.index('blank">') + 7:name.index('</a')]
        except ValueError:
            pass

        try:
            gender = tr[tr.index('Gender: <b>') + len('Gender: <b>'):
                        tr.index('Volleyball</b>', tr.index('Gender: <b>'))]
        except ValueError:
            if name:
                gender = ' '
            else:
                pass

        try:
            age = tr[tr.index('Team Ages:') + len('Team Ages:'):tr.index('</b>', tr.index('Team Ages:'))]
            age = age[4:]
        except ValueError:
            if name:
                age = ' '
            else:
                pass

        try:
            contact = tr[tr.index('Contact: <b>') + len('Contact: <b>'):tr.index('</b><br/>')]
            if len(contact) == 1:
                contact = ' '
        except ValueError:
            if name:
                contact = ' '
            else:
                pass

        try:
            state = tr[tr.index('State: <b>') + len('State: <b>'):tr.index('</b><br/>', tr.index('State: <b>'))]
            if '(' in state:
                state = state[:state.index('(')]
            if state == '':
                state = ' '
        except ValueError:
            if name:
                state = ' '
            else:
                pass

        try:
            city = tr[tr.index('City: <b>') + len('City: <b>'):tr.index('</b>', tr.index('City: <b>'))].title()
            if city == '':
                city = ' '
        except ValueError:
            if name:
                city = ' '
            else:
                pass

        try:
            club_page = tr[tr.index('</h3><a href="') + len('</h3><a href="'):tr.index('">More info')]
            club_url = r'http://www.usavolleyballclubs.com/' + club_page
            club_page = requests.get(club_url)
            club_page_source = club_page.content
            club_soup = BeautifulSoup(club_page_source, 'lxml')
            email = str(club_soup.find_all('a', {'target': '_top'})[0])
            email = email[email.index('mailto:') + len('mailto:'):email.index('cc=') - 5]
        except (ValueError, IndexError):
            if name:
                email = ' '
            else:
                pass

        try:
            website = tr[tr.index('<a href="') + len('<a href="'):tr.index('" target="', tr.index('<a href="'))]
        except ValueError:
            if name:
                website = ' '
            else:
                pass

        try:
            region = tr[tr.index('Club Volleyball Region:') + len('Club Volleyball Region:') + 4:tr.index('</b>', tr.index('Club Volleyball Region:'))]
            if '(' in region:
                region = region[region.index('(') + 1:region.index(')', region.index('('))]
            if region == '':
                region = 'No Region'
        except ValueError:
            if name:
                region = ' '
            else:
                pass

        current_row = [name, gender, age, contact, city, state, email, website, region]

        xl = pd.ExcelFile('test.xlsx')
        if region not in xl.sheet_names:
            wb.create_sheet('{}'.format(region))
            wb.save('test.xlsx')
            xl = pd.ExcelFile('test.xlsx')
            wb.worksheets[xl.sheet_names.index(region)].append(['Club Name', 'Gender', 'Age Group', 'Contact Name', 'City', 'State', 'Email', 'Website', 'Region'])
        wb.worksheets[xl.sheet_names.index(region)].append(current_row)
        wb.worksheets[0].append(current_row)
        wb.save('test.xlsx')

        index += 1
        print('\tEntry: ' + str(index))

        sleep(randint(2, 5))
